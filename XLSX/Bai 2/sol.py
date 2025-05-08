from openpyxl import load_workbook, Workbook

# Đọc file Excel gốc
input_file = 'grades_test_5.xlsx'
output_file = 'output.xlsx'

try:
    wb = load_workbook(input_file)
    ws = wb.active

    # Tạo workbook mới để ghi kết quả
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Kết quả"

    # Ghi tiêu đề
    ws_out.append(['Họ tên', 'Toán', 'Văn', 'Anh', 'Trung bình'])

    # Duyệt từ dòng 2 (bỏ qua header)
    for row in ws.iter_rows(min_row=2, values_only=True):
        ho_ten, toan, van, anh = row
        # Tính điểm trung bình
        tb = round((toan + van + anh) / 3, 2)
        # Ghi vào sheet mới
        ws_out.append([ho_ten, toan, van, anh, tb])

    # Lưu kết quả
    wb_out.save(output_file)
    print(f"✅ Đã ghi kết quả vào file '{output_file}'.")

except FileNotFoundError:
    print(f"❌ Không tìm thấy file '{input_file}'.")
except Exception as e:
    print(f"⚠️ Lỗi: {e}")
