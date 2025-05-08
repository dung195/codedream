from openpyxl import load_workbook, Workbook

input_file = 'employees_test_5.xlsx'
output_file = 'high_salary.xlsx'

try:
    # Đọc file đầu vào
    wb = load_workbook(input_file)
    ws = wb.active

    # Tạo file đầu ra
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Lương cao"

    # Ghi tiêu đề
    ws_out.append(["Tên", "Chức vụ", "Lương"])

    # Duyệt từng dòng dữ liệu (bỏ qua tiêu đề)
    for row in ws.iter_rows(min_row=2, values_only=True):
        ten, chuc_vu, luong = row
        if luong > 10_000_000:
            ws_out.append([ten, chuc_vu, luong])

    # Lưu file kết quả
    wb_out.save(output_file)
    print(f"✅ Đã lưu danh sách lương cao vào '{output_file}'.")

except FileNotFoundError:
    print(f"❌ Không tìm thấy file: {input_file}")
except Exception as e:
    print(f"⚠️ Lỗi: {e}")
