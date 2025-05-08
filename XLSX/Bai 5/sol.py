from openpyxl import load_workbook, Workbook

input_file = 'orders_test_5.xlsx'
output_file = 'orders_updated.xlsx'

try:
    # Đọc file đơn hàng
    wb = load_workbook(input_file)
    ws = wb.active

    # Duyệt qua các đơn hàng, bỏ qua tiêu đề
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ma_don, khach_hang, so_tien, trang_thai = row
        # Cập nhật trạng thái nếu số tiền = 0
        if so_tien == 0:
            ws.cell(row=row[0].row, column=4, value="Hủy")

    # Lưu kết quả vào file mới
    wb.save(output_file)
    print(f"✅ Đã lưu kết quả vào file '{output_file}'.")

except FileNotFoundError:
    print(f"❌ Không tìm thấy file: {input_file}")
except Exception as e:
    print(f"⚠️ Lỗi: {e}")
