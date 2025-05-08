from openpyxl import load_workbook, Workbook

input_file = 'products_test_5.xlsx'
output_file = 'categories_count.xlsx'

try:
    # Đọc file sản phẩm
    wb = load_workbook(input_file)
    ws = wb.active

    # Tạo dictionary để lưu số lượng sản phẩm theo danh mục
    category_count = {}

    # Duyệt qua tất cả các sản phẩm, bỏ qua tiêu đề
    for row in ws.iter_rows(min_row=2, values_only=True):
        ten_san_pham, danh_muc, gia = row
        if danh_muc in category_count:
            category_count[danh_muc] += 1
        else:
            category_count[danh_muc] = 1

    # Tạo file mới để lưu kết quả
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Số lượng sản phẩm"

    # Ghi tiêu đề
    ws_out.append(["Danh mục", "Số lượng sản phẩm"])

    # Ghi dữ liệu vào file output
    for danh_muc, count in category_count.items():
        ws_out.append([danh_muc, count])

    # Lưu file kết quả
    wb_out.save(output_file)
    print(f"✅ Đã lưu kết quả vào file '{output_file}'.")

except FileNotFoundError:
    print(f"❌ Không tìm thấy file: {input_file}")
except Exception as e:
    print(f"⚠️ Lỗi: {e}")
